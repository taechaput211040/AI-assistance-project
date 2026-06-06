# -*- coding: utf-8 -*-
"""
AI-assisted classification pipeline (v2).

Claude (the AI) read all 300 feedback texts, identified the distinct
semantic message-cores, and assigned Sentiment / Category / Priority / Owner
to each core. This script APPLIES those AI-made label decisions consistently
to every one of the 300 rows by matching each row to its semantic core.

Output:
  - exam-app/server/data/feedback.json   (consumed by the Nuxt dashboard)
  - deliverables/feedback_enriched.xlsx  (analysis-ready sheet)
"""
import json, os
from collections import Counter, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = "/sessions/hopeful-kind-hypatia/mnt/AI-assistance-project/player_feedback_300_dataset.xlsx"
JSON_OUT = "/sessions/hopeful-kind-hypatia/mnt/AI-assistance-project/exam-app/server/data/feedback.json"
XLSX_OUT = "/sessions/hopeful-kind-hypatia/mnt/AI-assistance-project/deliverables/feedback_analysis.xlsx"

# AI label table. Each tuple: (match_substring, category, sentiment, priority, owner, theme)
# Order matters: more specific cores are listed first.
P1, P2, P3 = "P1", "P2", "P3"
NEG, NEU, POS = "Negative", "Neutral", "Positive"
TABLE = [
    # --- Gacha / Monetization ---
    ("rate up", "Gacha / Monetization", NEG, P2, "Monetization", "กาชาโรลไม่ได้ rate up (ท้อ)"),
    ("Banner ใหม่ออกถี่", "Gacha / Monetization", NEG, P2, "Monetization", "Banner ออกถี่ เก็บเพชรไม่ทัน"),
    ("pity ชัดเจน", "Gacha / Monetization", NEG, P2, "Monetization", "ระบบ pity ไม่ชัดเจน"),
    ("ได้ตัวซ้ำบ่อย", "Gacha / Monetization", NEG, P2, "Monetization", "ตัวซ้ำ/แลกชิ้นส่วนไม่คุ้ม"),
    ("แพ็กเกจเติมเงินราคาสูง", "Gacha / Monetization", NEG, P2, "Monetization", "แพ็กเติมเงินแพงเทียบเพชร"),
    ("ราคา item ใน event shop สูง", "Gacha / Monetization", NEG, P2, "Monetization", "ราคาของใน event shop สูง"),
    ("pay to win", "Gacha / Monetization", NEG, P2, "Monetization", "รู้สึก pay-to-win มากขึ้น"),
    ("ถ้าไม่เติมเงินแทบตามคนอื่นไม่ทัน", "Gacha / Monetization", NEG, P2, "Monetization", "ไม่เติมเงินตามไม่ทัน (ทรัพยากรน้อย)"),
    # --- Gameplay / Balance ---
    ("one shot", "Gameplay / Balance", NEG, P2, "Game Design", "บอส one-shot ไม่แฟร์"),
    ("auto battle", "Gameplay / Balance", NEG, P2, "Game Design", "auto battle เลือกสกิลแย่"),
    ("meta แคบ", "Gameplay / Balance", NEG, P2, "Game Design", "PvP meta แคบ/ทีมโกง"),
    ("แรงเกินไป เจอใน arena", "Gameplay / Balance", NEG, P2, "Game Design", "ตัวละคร arena แรงเกิน (imbalance)"),
    ("ตัวละครสายฮีลดูอ่อน", "Gameplay / Balance", NEG, P2, "Game Design", "ตัวฮีลอ่อนเทียบดาเมจบอส"),
    ("balance ของอาวุธ", "Gameplay / Balance", NEG, P3, "Game Design", "อาวุธบาลานซ์ต่างกันเกิน"),
    ("พลังที่แนะนำในด่านไม่ตรง", "Gameplay / Balance", NEG, P3, "Game Design", "power แนะนำไม่ตรงความยากจริง"),
    ("ยากเกินไปสำหรับคนที่เพิ่งเริ่ม", "Gameplay / Balance", NEG, P2, "Game Design", "ด่านยากเกินสำหรับมือใหม่"),
    # --- Bug / Technical ---
    ("เกมแลคมากในด่านบอส", "Bug / Technical Issue", NEG, P1, "Engineering", "แลค/เฟรมตกในด่านบอส"),
    ("เด้งออกบ่อย", "Bug / Technical Issue", NEG, P1, "Engineering", "เด้งออกตอนเปลี่ยนแผนที่"),
    ("เกมค้างตอนเข้าหน้า", "Bug / Technical Issue", NEG, P1, "Engineering", "เกมค้างเมื่อเข้าบางหน้า"),
    ("ของไม่เข้า inventory", "Bug / Technical Issue", NEG, P1, "Engineering", "รับรางวัลแล้วของไม่เข้า"),
    ("เสียงเอฟเฟกต์หาย", "Bug / Technical Issue", NEG, P2, "Engineering", "เสียงเอฟเฟกต์หายต้อง restart"),
    ("โหลดแพตช์ช้า", "Bug / Technical Issue", NEG, P2, "Engineering", "โหลดแพตช์ช้าผิดปกติ"),
    ("แสดงผลเพี้ยน ปุ่มทับกัน", "Bug / Technical Issue", NEG, P2, "Engineering", "UI เพี้ยน/ปุ่มทับบนจอเล็ก"),
    ("แจ้งเตือนเควสต์ขึ้นซ้ำ", "Bug / Technical Issue", NEG, P3, "Engineering", "แจ้งเตือนเควสต์ค้าง/ซ้ำ"),
    # --- Account / Payment ---
    ("ล็อกอินด้วย Facebook ไม่ได้", "Account / Payment", NEG, P1, "Engineering", "ล็อกอิน Facebook ไม่ได้"),
    ("customer support ขอข้อมูลเยอะ", "Account / Payment", NEG, P1, "Player Support", "support ตอบช้า/ยังไม่ได้คำตอบ"),
    ("ตรวจประวัติการเติมเงิน", "Account / Payment", NEU, P3, "Player Support", "ขอช่องทางดูประวัติเติมเงิน"),
    # --- Reward / Economy ---
    ("ปลุกพลังหายากเกินไป", "Reward / Economy", NEG, P2, "Economy", "วัสดุปลุกพลังหายาก ฟาร์มไม่พอ"),
    ("ทองในเกมขาดตลอด", "Reward / Economy", NEG, P2, "Economy", "ทองไม่พออัปเกรด"),
    ("เควสต์รายสัปดาห์ให้แต้มไม่พอ", "Reward / Economy", NEG, P2, "Economy", "แต้มเควสต์รายสัปดาห์ไม่พอแลก"),
    ("ของรางวัลรายวันน้อย", "Reward / Economy", NEG, P2, "Economy", "รางวัลรายวันน้อยเทียบเวลาเล่น"),
    ("รางวัล login รอบนี้ดูธรรมดา", "Reward / Economy", NEG, P3, "Economy", "รางวัล login น่าเบื่อ"),
    ("stamina เยอะ แต่ของแลกในร้านไม่ค่อยคุ้ม", "Reward / Economy", NEG, P2, "Economy", "stamina แพงแต่ของแลกไม่คุ้ม"),
    # --- Event Feedback ---
    ("Event ใช้ stamina รวมกับด่านปกติ", "Event Feedback", NEG, P2, "LiveOps", "event ใช้ stamina ร่วมด่านปกติ"),
    ("ภารกิจ event บังคับเล่นหลายรอบ", "Event Feedback", NEG, P2, "LiveOps", "ภารกิจ event grind เยอะ"),
    ("เวลาจัด event สั้นเกินไป", "Event Feedback", NEG, P2, "LiveOps", "เวลา event สั้นเกินไป"),
    ("อันดับ event แข่งขันหนัก", "Event Feedback", NEG, P2, "LiveOps", "อันดับ event แข่งหนัก สู้ยาก"),
    ("ธีมน่ารักดี แต่ภารกิจซ้ำ", "Event Feedback", NEU, P3, "LiveOps", "ธีม event ดีแต่ภารกิจซ้ำ"),
    # --- Content Request ---
    ("ปฏิทิน event ล่วงหน้า", "Content Request", NEU, P3, "LiveOps", "ขอปฏิทิน event ล่วงหน้า"),
    ("เนื้อเรื่อง event สนุก แต่อยากให้มีเสียงพากย์", "Content Request", NEU, P3, "LiveOps", "ขอเสียงพากย์ในเนื้อเรื่อง event"),
    ("skin ฟรีจาก event", "Content Request", NEU, P3, "LiveOps", "ขอ skin ฟรีจาก event เพิ่ม"),
    ("preview animation", "Content Request", NEU, P3, "Product/UX", "ขอ preview ตัวละครก่อนสุ่ม"),
    ("subtitle ให้ครบ", "Content Request", NEU, P3, "Product/UX", "ขอภาษาไทย/subtitle ครบ"),
    ("preset ทีมหลายชุด", "Content Request", NEU, P3, "Product/UX", "ขอ preset ทีมหลายชุด"),
    ("claim all", "Content Request", NEU, P3, "Product/UX", "ขอปุ่ม claim all"),
    ("replay ดูการต่อสู้", "Content Request", NEU, P3, "Product/UX", "ขอระบบ replay การต่อสู้"),
    ("guild boss", "Content Request", NEU, P3, "Game Design", "ขอระบบ guild boss"),
    ("co-op เล่นกับเพื่อน", "Content Request", NEU, P3, "Game Design", "ขอโหมด co-op"),
    ("ตัวละครสายซัพพอร์ตใหม่", "Content Request", NEU, P3, "Game Design", "ขอตัวละครสายซัพพอร์ต"),
    ("stat เปรียบเทียบก่อนและหลังเปลี่ยนอุปกรณ์", "Content Request", NEU, P3, "Product/UX", "ขอเทียบ stat ก่อน/หลังสวมอุปกรณ์"),
    # --- UX / UI ---
    ("ตัวหนังสือในหน้ารายละเอียดสกิลเล็ก", "UX / UI", NEG, P3, "Product/UX", "ฟอนต์รายละเอียดสกิลเล็ก"),
    ("แจ้งเตือนสีแดงเยอะเกินไป", "UX / UI", NEG, P3, "Product/UX", "badge แจ้งเตือนสีแดงเยอะ"),
    ("inventory กรองของได้ไม่ละเอียด", "UX / UI", NEG, P3, "Product/UX", "ตัวกรอง inventory หยาบ"),
    ("หาเจอยาก ต้องกดหลายขั้นตอน", "UX / UI", NEG, P3, "Product/UX", "เมนูเข้าถึงยาก หลายขั้นตอน"),
    ("ปุ่มย้อนกลับบางหน้าพาออกไปหน้าแรก", "UX / UI", NEG, P3, "Product/UX", "ปุ่มย้อนกลับเด้งหน้าแรก"),
    ("Tutorial ช่วงแรกเร็วไป", "UX / UI", NEG, P2, "Product/UX", "tutorial เร็ว/ข้อมูลแน่นเกิน"),
    ("หน้า shop อธิบายสิทธิ์ไม่ชัด", "UX / UI", NEU, P3, "Product/UX", "shop อธิบายสิทธิ์ไม่ชัด"),
    # --- Positive Feedback ---
    ("ชอบตัวละครใหม่มาก", "Positive Feedback", POS, P3, "Community", "ชมตัวละคร/ดีไซน์ใหม่"),
    ("ภาพสวยขึ้นมาก", "Positive Feedback", POS, P3, "Community", "ชมกราฟิก/ฉากต่อสู้ใหม่"),
    ("UI ใหม่ดูสะอาดขึ้น", "Positive Feedback", POS, P3, "Community", "ชม UI ใหม่"),
    ("mini game ใน event นี้ เล่นง่าย", "Positive Feedback", POS, P3, "Community", "ชม mini game ของ event"),
    ("กลับมาเล่นทุกวันอีกครั้ง", "Positive Feedback", POS, P3, "Community", "event ดึงให้กลับมาเล่น"),
    ("ของแจกครบรอบดีเกินคาด", "Positive Feedback", POS, P3, "Community", "ชมของแจกครบรอบ"),
    ("เพลงในแผนที่ใหม่ดีมาก", "Positive Feedback", POS, P3, "Community", "ชมเพลง/แผนที่ใหม่"),
    ("ตอบปัญหาใน community เร็ว", "Positive Feedback", POS, P3, "Community", "ชมทีม support/community"),
]

PRIO_W = {"P1": 3, "P2": 2, "P3": 1}
PRIO_TH = {"P1": "เร่งด่วน", "P2": "สำคัญ", "P3": "ปรับปรุงย่อย"}

def make_summary(theme, sent, prio, owner):
    tag = PRIO_TH[prio]
    if sent == POS:
        return f"คำชม: {theme} — ส่งทีม {owner} นำไปต่อยอด/สื่อสาร"
    if sent == NEU:
        return f"ข้อเสนอแนะ ({tag}): {theme} — ส่งทีม {owner} พิจารณา"
    return f"ปัญหา ({tag}): {theme} — ทีม {owner} ควรแก้ไข"

def classify(text):
    for sub, cat, sent, prio, owner, theme in TABLE:
        if sub in text:
            return cat, sent, prio, owner, theme
    return None

def main():
    wb = openpyxl.load_workbook(SRC, read_only=True)
    rows = list(wb["Feedback_Raw"].iter_rows(values_only=True))
    header = rows[0]
    records, unmatched = [], []
    for r in rows[1:]:
        d = dict(zip(header, r))
        text = (d.get("player_feedback") or "").strip()
        res = classify(text)
        if res is None:
            unmatched.append((d.get("feedback_id"), text))
            res = ("UX / UI", NEU, P3, "Product/UX", "อื่น ๆ")
        cat, sent, prio, owner, theme = res
        dt = d.get("date")
        records.append({
            "id": d.get("feedback_id"),
            "date": dt.isoformat() if hasattr(dt, "isoformat") else str(dt),
            "source": d.get("source"), "player_id": d.get("player_id"),
            "segment": d.get("player_segment"), "platform": d.get("platform"),
            "version": d.get("game_version"), "area": d.get("game_area_hint"),
            "text": text, "category": cat, "sentiment": sent,
            "priority": prio, "owner": owner, "theme": theme,
            "ai_summary": make_summary(theme, sent, prio, owner),
            "priority_score": PRIO_W[prio] * (1 if sent == NEG else 0.4 if sent == NEU else 0),
        })

    if unmatched:
        print("UNMATCHED:", len(unmatched))
        for u in unmatched[:20]:
            print("  ", u)

    # ---- aggregates ----
    def dist(key):
        return dict(Counter(r[key] for r in records))

    total = len(records)
    sent_tot = Counter(r["sentiment"] for r in records)
    prio_tot = Counter(r["priority"] for r in records)

    sent_by_cat = defaultdict(lambda: {POS: 0, NEU: 0, NEG: 0})
    for r in records:
        sent_by_cat[r["category"]][r["sentiment"]] += 1

    # themes ranked by impact
    th = defaultdict(lambda: {"count": 0, "neg": 0, "p1": 0, "impact": 0.0, "category": "", "owner": ""})
    for r in records:
        t = th[r["theme"]]
        t["count"] += 1; t["category"] = r["category"]; t["owner"] = r["owner"]
        if r["sentiment"] == NEG: t["neg"] += 1
        if r["priority"] == P1: t["p1"] += 1
        t["impact"] += r["priority_score"]
    themes = []
    for name, v in th.items():
        themes.append({"theme": name, "category": v["category"], "owner": v["owner"],
                       "count": v["count"], "negative": v["neg"], "p1": v["p1"],
                       "neg_rate": round(v["neg"]/v["count"]*100),
                       "impact": round(v["impact"])})
    themes.sort(key=lambda x: x["impact"], reverse=True)

    # owner workload
    owner_agg = defaultdict(lambda: {"count": 0, "neg": 0, "p1": 0})
    for r in records:
        o = owner_agg[r["owner"]]
        o["count"] += 1
        if r["sentiment"] == NEG: o["neg"] += 1
        if r["priority"] == P1: o["p1"] += 1
    owners = [{"owner": k, **v, "neg_rate": round(v["neg"]/v["count"]*100)} for k, v in owner_agg.items()]
    owners.sort(key=lambda x: (x["p1"], x["neg"]), reverse=True)

    # category summary + actions
    ACTIONS = {
        "Bug / Technical Issue": "ตั้ง hotfix sprint แก้ crash/lag/ของไม่เข้าตามลำดับความถี่ + เพิ่ม crash logging",
        "Account / Payment": "เร่งเคสล็อกอิน/เติมเงินรายกรณี ตั้ง SLA ตอบ support ภายใน 24 ชม.",
        "Gacha / Monetization": "ทำหน้า pity/อัตราดรอปให้โปร่งใส ทบทวนราคาแพ็ก/ความถี่ banner ลดภาพ pay-to-win",
        "Reward / Economy": "ปรับ drop rate วัสดุปลุกพลัง/ทอง/stamina และทบทวนคุณค่ารางวัล login/รายวัน",
        "Gameplay / Balance": "ทบทวนบาลานซ์บอส/ตัว arena/ตัวฮีล แก้ AI auto-battle และเส้นความยากมือใหม่",
        "Event Feedback": "ลดการ grind ของ event, แยก stamina, ยืดเวลา และเพิ่มปฏิทินล่วงหน้า",
        "UX / UI": "ปรับ tutorial ให้ข้ามได้, แก้ปุ่มย้อนกลับ, เพิ่มตัวกรอง inventory, จัดลำดับ badge แจ้งเตือน",
        "Content Request": "รวบรวมเป็น backlog จัดลำดับตามจำนวนโหวต (replay, co-op, claim all, preset ฯลฯ)",
        "Positive Feedback": "ขยายผลสิ่งที่ผู้เล่นชอบ (ธีม event, ดีไซน์, UI ใหม่) และใช้สื่อสารการตลาด",
    }
    cat_summary = []
    for cat, sc in sent_by_cat.items():
        cnt = sum(sc.values())
        cat_summary.append({"category": cat, "count": cnt, "positive": sc[POS],
                            "neutral": sc[NEU], "negative": sc[NEG],
                            "neg_rate": round(sc[NEG]/cnt*100), "action": ACTIONS.get(cat, "")})
    cat_summary.sort(key=lambda x: x["negative"], reverse=True)

    # time trend
    trend = defaultdict(lambda: {POS: 0, NEU: 0, NEG: 0})
    for r in records:
        trend[r["date"][:10]][r["sentiment"]] += 1
    trend_list = [{"date": d, **v} for d, v in sorted(trend.items())]

    meta = {
        "total": total,
        "sentiment": {"Positive": sent_tot[POS], "Neutral": sent_tot[NEU], "Negative": sent_tot[NEG]},
        "priority": {"P1": prio_tot[P1], "P2": prio_tot[P2], "P3": prio_tot[P3]},
        "neg_rate": round(sent_tot[NEG]/total*100),
        "date_min": min(r["date"][:10] for r in records),
        "date_max": max(r["date"][:10] for r in records),
        "by_category": dist("category"), "by_source": dist("source"),
        "by_segment": dist("segment"), "by_platform": dist("platform"),
        "by_version": dist("version"), "by_owner": dist("owner"),
        "sentiment_by_category": {k: dict(v) for k, v in sent_by_cat.items()},
        "category_summary": cat_summary, "owners": owners,
        "top_themes": themes, "trend": trend_list,
        "method": "AI semantic classification (Claude) applied consistently across templated rows",
    }

    os.makedirs(os.path.dirname(JSON_OUT), exist_ok=True)
    with open(JSON_OUT, "w", encoding="utf-8") as f:
        json.dump({"meta": meta, "records": records}, f, ensure_ascii=False, indent=1)

    # ---- enriched xlsx ----
    build_xlsx(records, meta)

    print("records:", total)
    print("sentiment:", dict(sent_tot))
    print("priority:", dict(prio_tot))
    print("by_category:", json.dumps(meta["by_category"], ensure_ascii=False))
    print("by_owner:", json.dumps(meta["by_owner"], ensure_ascii=False))
    print("top8 themes:")
    for t in themes[:8]:
        print(f"   {t['theme']} | n={t['count']} neg%={t['neg_rate']} P1={t['p1']} impact={t['impact']} -> {t['owner']}")


def build_xlsx(records, meta):
    os.makedirs(os.path.dirname(XLSX_OUT), exist_ok=True)
    wb = openpyxl.Workbook()
    head_fill = PatternFill("solid", fgColor="1F2A44")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D0D7E2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sent_fill = {"Negative": "FDE2E4", "Neutral": "FFF3CD", "Positive": "D7F5E3"}
    prio_fill = {"P1": "F8B4B4", "P2": "FCD9A8", "P3": "DCE3F0"}

    # Sheet 1: cleaned + analysed feedback
    ws = wb.active; ws.title = "Feedback_Analysis"
    cols = ["feedback_id", "date", "source", "player_segment", "platform", "game_version",
            "game_area_hint", "player_feedback", "Category", "Sentiment",
            "Priority", "AI_Summary", "Suggested_Owner", "Theme"]
    ws.append(cols)
    for c in range(1, len(cols)+1):
        cell = ws.cell(1, c); cell.fill = head_fill; cell.font = head_font
        cell.alignment = Alignment(vertical="center", horizontal="center"); cell.border = border
    for r in records:
        ws.append([r["id"], r["date"][:16].replace("T", " "), r["source"], r["segment"],
                   r["platform"], r["version"], r["area"], r["text"], r["category"],
                   r["sentiment"], r["priority"], r["ai_summary"], r["owner"], r["theme"]])
        row = ws.max_row
        ws.cell(row, 10).fill = PatternFill("solid", fgColor=sent_fill[r["sentiment"]])
        ws.cell(row, 11).fill = PatternFill("solid", fgColor=prio_fill[r["priority"]])
        ws.cell(row, 8).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 12).alignment = Alignment(wrap_text=True, vertical="top")
    widths = [10, 16, 16, 14, 11, 10, 15, 52, 21, 10, 8, 50, 13, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(cols))}{ws.max_row}"

    # Sheet 2: category summary
    ws2 = wb.create_sheet("Category_Summary")
    ws2.append(["Category", "Total", "Negative", "Neutral", "Positive", "Neg %", "Recommended Action"])
    for c in range(1, 8):
        cell = ws2.cell(1, c); cell.fill = head_fill; cell.font = head_font; cell.border = border
    for c in meta["category_summary"]:
        ws2.append([c["category"], c["count"], c["negative"], c["neutral"], c["positive"],
                    f'{c["neg_rate"]}%', c["action"]])
    for i, w in enumerate([24, 8, 10, 9, 9, 8, 70], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for row in ws2.iter_rows(min_row=2):
        row[6].alignment = Alignment(wrap_text=True, vertical="top")

    # Sheet 3: top themes
    ws3 = wb.create_sheet("Top_Issues")
    ws3.append(["Rank", "Theme", "Category", "Owner", "Count", "Negative", "Neg %", "P1", "Impact"])
    for c in range(1, 10):
        cell = ws3.cell(1, c); cell.fill = head_fill; cell.font = head_font; cell.border = border
    for i, t in enumerate(meta["top_themes"], 1):
        ws3.append([i, t["theme"], t["category"], t["owner"], t["count"], t["negative"],
                    f'{t["neg_rate"]}%', t["p1"], t["impact"]])
    for i, w in enumerate([6, 36, 22, 14, 8, 10, 8, 6, 9], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Sheet 4: owner workload
    ws4 = wb.create_sheet("Owner_Workload")
    ws4.append(["Owner (Team)", "Total", "Negative", "Neg %", "P1 (urgent)"])
    for c in range(1, 6):
        cell = ws4.cell(1, c); cell.fill = head_fill; cell.font = head_font; cell.border = border
    for o in meta["owners"]:
        ws4.append([o["owner"], o["count"], o["neg"], f'{o["neg_rate"]}%', o["p1"]])
    for i, w in enumerate([18, 8, 10, 8, 12], 1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(XLSX_OUT)


if __name__ == "__main__":
    main()
